# -*-coding:utf-8 -*- #
# ---------------------------------------------------------------------------
# ProjectName:   AutoTest
# FileName:      read_excel.py
# Author:       zdq
# Datetime:     2023/12/4 10:51
# Description:
# 命名规范：文件名全小写+下划线，类名大驼峰，方法和变量小写+下划线连接，
# 常量大写，变量和常量用名词，方法用动词
#
import os

import openpyxl

from common.read_ini import ReadIni
from common.read_json import read_json
from config.settings import *


class ReadExcel:
    def __init__(self, table_name, excel_name, case_data_name, expect_data_name, sql_data_name=None):
        self.read_ini = ReadIni()
        excel_path = os.path.join(self.read_ini.data_config_path, excel_name)
        case_data_path = os.path.join(self.read_ini.data_config_path, case_data_name)
        expect_data_path = os.path.join(self.read_ini.data_config_path, expect_data_name)
        if sql_data_name is not None:
            sql_data_path = os.path.join(self.read_ini.data_config_path, sql_data_name)
            self.sql_data_dict = read_json(sql_data_path)
        self.case_data_dict = read_json(case_data_path)
        self.expect_data_dict = read_json(expect_data_path)

        self.wb = openpyxl.load_workbook(excel_path)
        self.ws = self.wb[table_name]

    def __get_cell_value(self, column, row):
        value = self.ws[column + str(row)].value
        return value

    def get_module_name(self, row):
        return self.__get_cell_value(MODULE, row)

    def get_api_name(self, row):
        return self.__get_cell_value(API, row)

    def get_title(self, row):
        return self.__get_cell_value(TITLE, row)

    def get_level(self, row):
        return self.__get_cell_value(LEVEL, row)

    def get_case_url(self, row):
        api = self.__get_cell_value(URL, row)
        case_url = self.read_ini.get_host(SERVER_HOST)
        return case_url + api

    def get_case_method(self, row):
        return self.__get_cell_value(METHOD, row)

    def get_case_mime(self, row):
        return self.__get_cell_value(MINI, row)

    def get_case_data(self, row):
        case_data_key = self.__get_cell_value(CASE_DATA, row)
        module_name = self.get_module_name(row)
        api_name = self.get_api_name(row)
        return self.case_data_dict[module_name][api_name][case_data_key]

    def get_expect_data(self, row):
        expect_data_key = self.__get_cell_value(EXPECT_DATA, row)
        module_name = self.get_module_name(row)
        api_name = self.get_api_name(row)
        return self.expect_data_dict[module_name][api_name][expect_data_key]

    def get_sql_type(self, row):
        return self.__get_cell_value(SQL_TYPE, row)

    def get_sql_sentence(self, row):
        sql_data_key = self.__get_cell_value(SQL_DATA, row)
        module_name = self.get_module_name(row)
        api_name = self.get_api_name(row)
        if sql_data_key:
            return self.expect_data_dict[module_name][api_name][sql_data_key]
        else:
            return None

    def get_sql_update_key(self, row):
        return self.__get_cell_value(UPDATE_KEY, row)

    def get_data(self):
        list_data = []
        for row in range(2, self.ws.max_row+1):
            module_name = self.get_module_name(row)
            api_name = self.get_api_name(row)
            title = self.get_title(row)
            level = self.get_level(row)
            case_method = self.get_case_method(row)
            case_url = self.get_case_url(row)
            case_mini = self.get_case_mime(row)
            case_data = self.get_case_data(row)
            case_expect = self.get_expect_data(row)
            sql_type = self.get_sql_type(row)
            sql_data = self.get_sql_sentence(row)
            update_key = self.get_sql_update_key(row)
            list_data.append(
                [module_name, api_name, title, level, case_method, case_url, case_mini, case_data, case_expect, sql_type, sql_data, update_key]
            )
        else:
            return list_data


if __name__ == '__main__':
    print(ReadExcel("登录", "../config/zdq/APIAutoTest.xlsx", "../config/zdq/case_data.json", "../config/zdq/expect_data.json").get_data())