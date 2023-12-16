# -*-coding:utf-8 -*- #
# ---------------------------------------------------------------------------
# ProjectName:   i2share_selenium
# FileName:      read_excel.py
# Author:       zdq
# Datetime:     2023/12/13 15:34
# Description:
# 命名规范：文件名全小写+下划线，类名大驼峰，方法和变量小写+下划线连接，
# 常量大写，变量和常量用名词，方法用动词
#
import openpyxl

from common.read_ini import ReadIni
from common.read_json import read_json


class ReadExcel:
    def __init__(self, table_name="Sheet1"):
        readini = ReadIni()
        excel_path = readini.get_file_path("excel")
        case_data_path = readini.get_file_path("case")

        self.case_data_dict = read_json(case_data_path)

        wb = openpyxl.load_workbook(excel_path)
        self.ws = wb[table_name]

    def __get_cell_value(self, column, row):
        value = self.ws[str(column) + str(row)].value
        if value is None:
            return None
        elif value.strip():
            return value.strip()

    def module_name(self, row):
        return self.__get_cell_value("b", row)

    def func_name(self, row):
        return self.__get_cell_value("c", row)

    def case_data(self, row):
        case_data_key = self.__get_cell_value("f", row)

        if case_data_key is None:
            return None
        elif case_data_key:
            module_name = self.module_name(row)
            func_name = self.func_name(row)
            return self.case_data_dict[module_name][func_name][case_data_key]

    def expect_data(self, row):
        return self.__get_cell_value("g", row)

    def sql_data(self, row):
        return self.__get_cell_value("h", row)

    def get_data(self):
        list_data = []

        for row in range(2, self.ws.max_row + 1):
            case_data = self.case_data(row)
            expect_data = self.expect_data(row)
            sql_data = self.sql_data(row)
            list_data.append([case_data, expect_data, sql_data])
        else:
            return list_data


if __name__ == '__main__':
    da = ReadExcel().get_data()
    print(da[0])