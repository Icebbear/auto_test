# -*-coding:utf-8 -*- #
# ---------------------------------------------------------------------------
# ProjectName:   test_57_unittest
# FileName:      read_excel.py
# Author:       ice bear
# Datetime:     2023/12/2 11:43
# Description:
# 
# ---------------------------------------------------------------------------
import openpyxl

from common.read_ini import ReadIni
from common.read_json import read_json


class ReadExcel:
    def __init__(self, table_name="Sheet1"):
        read_ini = ReadIni()
        excal_path = read_ini.get_file_path("excel")

        case_data_path = read_ini.get_file_path("case")
        self.case_data_dict = read_json(case_data_path)

        wb = openpyxl.load_workbook(excal_path)
        self.ws = wb[table_name]

    def __get_cell_value(self, column, row) -> str:
        value = self.ws[str(column) + str(row)].value

        if value is None:
            return None
        elif value.strip():
            return value.strip()

    def module_name(self, row):
        return self.__get_cell_value("b", row)

    def func_name(self, row):
        return self.__get_cell_value("c", row)

    def case_data(self, row):
        case_data_key = self.__get_cell_value("f", row)

        if case_data_key is None:
            return None
        elif case_data_key:
            module_name = self.module_name(row)
            func_name = self.func_name(row)
            return self.case_data_dict[module_name][func_name][case_data_key]

    def expect_data(self, row):
        return self.__get_cell_value("g", row)

    def sql_data(self, row):
        return self.__get_cell_value("h", row)

    def get_data(self):
        list_data = []

        for row in range(2, self.ws.max_row+1):
            case_data = self.case_data(row)
            expect_data = self.expect_data(row)
            sql_data = self.sql_data(row)
            list_data.append([case_data, expect_data, sql_data])
        else:
            return list_data


if __name__ == '__main__':
    read_excel = ReadExcel()
    print(read_excel.get_data()[0])