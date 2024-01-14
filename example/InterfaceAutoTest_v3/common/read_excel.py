# -*-coding:utf-8 -*- #
# ---------------------------------------------------------------------------
# ProjectName:   test_57
# FileName:      read_excel.py
# Author:       ice bear
# Datetime:     2023/11/3 14:31
# Description:
# 
# ---------------------------------------------------------------------------
import os

import openpyxl
from InterfaceAutoTest_v3.data_config.settings import *
from InterfaceAutoTest_v3.common import log
from InterfaceAutoTest_v3.common.read_ini import ReadIni
from InterfaceAutoTest_v3.common.read_json import read_json


class ReadExcel:
    def __init__(self, table_name, excel_name, case_data_name, expect_data_name, sql_data_name=None):
        """获取数据配置层中除了ini文件以外的所有文件的路径，再获取excel的工作表，再读取所有的json文件"""
        self.read_ini = ReadIni()

        # 单数据文件--单用户
        """
        excel_path = self.read_ini.get_file_path(EXCEL)
        case_data_path = self.read_ini.get_file_path(CASE)
        expect_data_path = self.read_ini.get_file_path(EXPECT)
        sql_data_path = self.read_ini.get_file_path(SQL)
        # 获取工作表名称
        # table_name = self.read_ini.get_table_name(TABLE_NAME)
        """
        # 多数据文件--多用户
        excel_path = os.path.join(self.read_ini.data_config_path, excel_name)
        case_data_path = os.path.join(self.read_ini.data_config_path, case_data_name)
        expect_data_path = os.path.join(self.read_ini.data_config_path, expect_data_name)
        if sql_data_name is not None:
            sql_data_path = os.path.join(self.read_ini.data_config_path, sql_data_name)

        # table_name = table_name
        try:
            self.wb = openpyxl.load_workbook(excel_path)
        except:
            log.error("excel文件的路径错误")

        try:
           self.ws = self.wb[table_name]
        except:
            log.error("获取工作表失败，请察看工作表名称是否配置正确！！！")
            raise KeyError("获取工作表失败，请察看工作表名称是否配置正确！！！")

        self.case_data_dict = read_json(case_data_path)
        self.expect_data_dict = read_json(expect_data_path)
        self.sql_data_dict = read_json(sql_data_path)

    def __get_cell_value(self, column: str, row: int) -> str:
        """获取指定单元格数据"""
        try:
            value = self.ws[column+str(row)].value
            if value is None:
                return None
            elif value.strip():
                return value.strip()
        except:
            log.error("获取指定单元格数据失败，请察看输入的列号和行号是否正确！！")
            raise KeyError("获取指定单元格数据失败，请察看输入的列号和行号是否正确！！")

    def module_name(self, row):
        """根据行号，获取模块名称"""
        return self.__get_cell_value(MODULE, row)

    def api_name(self, row):
        """根据行号，获取接口名称"""
        return self.__get_cell_value(API, row)

    def title(self, row):
        """根据行号，获取用例标题"""
        return self.__get_cell_value(TITLE, row)

    def level(self, row):
        """根据行号，获取用例等级"""
        return self.__get_cell_value(LEVEL, row)

    def case_url(self, row):
        """根据行号，获取请求的url"""
        return self.read_ini.get_host(URL_HOST_CUSTOM) + self.__get_cell_value(URL, row)

    def case_method(self, row):
        """根据行号，获取请求方法"""
        return self.__get_cell_value(METHOD, row)

    def case_mime(self, row):
        """根据行号，获取媒体类型"""
        mime = self.__get_cell_value(MIME, row)

        if mime:
            return mime.lower()

    def case_data(self, row):
        """根据行号，获取用例数据"""
        case_data_key = self.__get_cell_value(CASE_DATA, row)

        if case_data_key:
            module_name = self.module_name(row)
            api_name = self.api_name(row)
            try:
                return self.case_data_dict[module_name][api_name][case_data_key]
            except:
                log.error("请察看是否配置了用例数据！！！")
                raise KeyError("请察看是否配置了用例数据！！！")

    def expect_data(self, row):
        """根据行号，获取期望数据"""
        expect_data_key = self.__get_cell_value(EXPECT_DATA, row)
        module_name = self.module_name(row)
        api_name = self.api_name(row)
        try:
            return self.expect_data_dict[module_name][api_name][expect_data_key]
        except:
            log.error("请察看是否配置了期望数据！！！")
            raise KeyError("请察看是否配置了期望数据！！！")

    def sql_type(self, row):
        """根据行号，获取sql语句类型"""
        sql_type_value = self.__get_cell_value(SQL_TYPE, row)

        if sql_type_value:
            return sql_type_value.lower()

    def sql_data(self, row):
        """根据行号，获取sql语句"""
        sql_data_key = self.__get_cell_value(SQL_DATA, row)

        if sql_data_key:
            module_name = self.module_name(row)
            api_name = self.api_name(row)
            try:
                return self.sql_data_dict[module_name][api_name][sql_data_key]
            except:
                log.error("请察看是否配置了sql数据！！！")
                raise KeyError("请察看是否配置了sql数据！！！")

    def update_key(self, row):
        """根据行号，获取更新的key"""
        return self.__get_cell_value(UPDATE_KEY, row)

    def get_data(self):
        """将测试数据存放在一个二维列表中"""
        list_data = []
        try:
            for row in range(2, self.ws.max_row + 1):
                # module_name
                module_name = self.module_name(row)
                # api
                api_name = self.api_name(row)
                # title
                title = self.title(row)
                # level
                level = self.level(row)
                # url
                case_url = self.case_url(row)
                # method
                case_method = self.case_method(row)
                # mime
                case_mime = self.case_mime(row)
                # case_data
                case_data = self.case_data(row)
                # expect_data
                expect_data = self.expect_data(row)
                # sql_type
                sql_type = self.sql_type(row)
                # sql_data
                sql_data = self.sql_data(row)
                # update_key
                update_key = self.update_key(row)
                list_data.append(
                    [module_name, api_name, title, level, case_url, case_method, case_mime, case_data, expect_data, sql_type, sql_data, update_key])
            else:
                return list_data
        except:
            log.error("请察看excel中是否存在空行，如果存在，请删除空行！！")
            raise TypeError("请察看excel中是否存在空行，如果存在，请删除空行！！")


if __name__ == '__main__':
    # excel = ReadExcel()
    # print(excel.get_data())
    print(ReadExcel("BPM-laohuang", "laohuang/APIAutoTest.xlsx", "laohuang/case_data.json", "laohuang/expect_data.json",
                    "laohuang/sql_data.json").get_data())